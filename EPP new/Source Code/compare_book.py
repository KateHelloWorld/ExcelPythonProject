from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tkinter import *
from tkinter import ttk, messagebox
from datetime import datetime, date, timedelta
from os.path import exists

from settings import Settings
from create_folder import CreateFolder


class CompareBook:

	"""
		Object creates analysis file and saves it in date folder. Also it stores usefull information related to the analysis folder.
	"""
	def __init__(self):
		self.settings = Settings()
		cf = CreateFolder()
		self.directory = cf.pnl_dir

		self.bookname = ""

	def check_files(self):
		bookname = self.bookname
		book_exists = exists(self.directory+"/"+bookname+".xlsx")
		if not book_exists:
			messagebox.showwarning("File "+bookname+".xlsx not found","File "+self.directory+"/"+" does not exists")
			return False
		external_exists = exists(self.directory+"/external.xlsx")
		if not external_exists:
			messagebox.showwarning("File external.xlsx not found","File "+self.directory+"/external.xlsx"+" does not exists")
			return False
		return True

	def get_system_names(self, root):
		bookname = self.bookname
		root.geometry("530x230")
		window = Frame(root, relief = 'sunken', bd=1, bg="#FFFFFF")
		window.pack(fill='both', expand=True, padx=10, pady=10)

		def set_systems():
			self.settings.accr_sys_name = drop_a.get()
			self.settings.fund_sys_name = drop_f.get()
			self._create_analisys_file()
			quit()

		def quit():
			root.destroy()

		clicked_a = StringVar()
		clicked_f = StringVar()

		message1 = Label(window, text="Please choose system1:", bg="#FFFFFF")
		message1.grid(row=1, column=0, padx=5)
		message2 = Label(window, text="Please choose system2:", bg="#FFFFFF")
		message2.grid(row=2, column=0, padx=5)

		systems = self._get_system_names()

		drop_f = ttk.Combobox(window, width=30, textvariable=clicked_f)
		drop_f['values'] = systems
		drop_f.grid(row=2, column=1, padx=5)
		drop_a = ttk.Combobox(window, width=30, textvariable=clicked_a)
		drop_a['values'] = systems 
		drop_a.grid(row=1, column=1, pady=20, padx=5)

		drop_a.set("System1")
		drop_f.set("System2")

		Checkbutton1 = BooleanVar()
		ch_button1 = Checkbutton(window, text="Add currency information", variable=Checkbutton1, onvalue=True, offvalue=False, bg="#FFFFFF", pady=20)
		ch_button1.select()
		ch_button1.grid(row=3, column=0)

		button1 = Button(window, text="Submit", command=set_systems, width=10, bg="#DDDDDD", activebackground="#BBBBBB")
		button1.grid(row=4, column=0)
		button2 = Button(window, text="Quit", command=quit, width=10, bg="#DDDDDD", activebackground="#BBBBBB")
		button2.grid(row=4, column=1)

		window.mainloop()

	def _load_data_from_external(self, analisys_data):
		bookname = self.bookname
		settings = self.settings
		external = load_workbook(self.directory+"/external.xlsx")
		external_sheet = external["Sheet1"]
		all_checked = False
		counter = 2
		counter_analisys = 2

		while not all_checked and counter<70000:
			if str(external_sheet[settings.external_id_col+str(counter)]) == 'None':
				all_checked = true
			else:
				if str(external_sheet[settings.external_book_col+str(counter)].value) == str(bookname):
					if not external_sheet[settings.external_mult_col+str(counter)].value == 0 and not str(external_sheet[settings.external_mult_col+str(counter)].value) == 'None':
						analisys_data[settings.data_external_id_col+str(counter)] = external_sheet[settings.external_id_col+str(counter)].value 
						mult = float(str(external_sheet[settings.external_mult_col+str(counter)].value))
						analisys_data[settings.data_external_a_col+str(counter)] = float(str(external_sheet[settings.external_accr_col+str(counter)].value)) * mult
						analisys_data[settings.data_external_f_col+str(counter)] = float(str(external_sheet[settings.external_fund_col+str(counter)].value)) * mult
				counter = counter+1

		print("Data from external file loaded successfully!")

	def _load_data_from_internal(self, analisys_data, analisis_output):
		bookname = self.bookname
		book = load_workbook(self.directory+"/"+bookname+".xlsx")
		book_sheet = book.active
		
		settings = self.settings
		tradeid_list = []
		accr_dict = {}
		fund_dict = {}
		accr_fx = {}
		fund_fx = {}
		accr_flag = False
		fund_flag = False
		accr_ccy_flag = False
		fund_ccy_flag = False

		#Load trade id and values for accural and funding in dictionaries
		curr = 2
		while curr<70000:
			if str(book_sheet[settings.system_col+str(curr)].value) == settings.accr_sys_name:
				accr_flag = True
				fund_flag = False
			elif str(book_sheet[settings.system_col+str(curr)].value) == settings.fund_sys_name:
				accr_flag = False
				fund_flag = True

			if accr_flag:
				accr_dict[book_sheet[settings.book_id_col+str(curr)].value] = {}
				accr_dict[book_sheet[settings.book_id_col+str(curr)].value]["EUR"] = book_sheet[settings.accr_EUR_col+str(curr)].value
				accr_dict[book_sheet[settings.book_id_col+str(curr)].value]["CCY"] = book_sheet[settings.accr_ccy_col+str(curr)].value 
				accr_dict[book_sheet[settings.book_id_col+str(curr)].value]["ccy"] = book_sheet[settings.accr_CCY_col+str(curr)].value

			elif fund_flag:
				fund_dict[book_sheet[settings.book_id_col+str(curr)].value] = {}
				fund_dict[book_sheet[settings.book_id_col+str(curr)].value]["EUR"] = book_sheet[settings.fund_EUR_col+str(curr)].value
				fund_dict[book_sheet[settings.book_id_col+str(curr)].value]["CCY"] = book_sheet[settings.fund_ccy_col+str(curr)].value 
				fund_dict[book_sheet[settings.book_id_col+str(curr)].value]["ccy"] = book_sheet[settings.fund_CCY_col+str(curr)].value


			if not str(book_sheet[settings.book_id_col+str(curr)].value) == 'None':
				tradeid_list.append(book_sheet[settings.book_id_col+str(curr)].value)

			curr = curr+1

		# Remove duplicates
		tradeid_list = list(set(tradeid_list))

		#insert trade id into output sheet
		counter_analisys = 2 
		for key in tradeid_list:
			analisis_output[settings.output_id_col+str(counter_analisys)] = key
			counter_analisys = counter_analisys+1

		#insert trade id and accrual values into analisys data
		#insert trade id and accrual values into analisys data
		counter_analisys = 2 
		for key in accr_dict:
			analisys_data[settings.data_accr_id_col+str(counter_analisys)] = key
			analisys_data[settings.data_accr_EUR_col+str(counter_analisys)] = accr_dict[key]["EUR"]
			analisys_data[settings.data_accr_ccy_col+str(counter_analisys)] = accr_dict[key]["ccy"]
			analisys_data[settings.data_accr_CCY_col+str(counter_analisys)] = accr_dict[key]["CCY"]
			counter_analisys = counter_analisys+1

		#insert trade id and funding values into analisys data
		counter_analisys = 2 
		for key in fund_dict:
			analisys_data[settings.data_fund_id_col+str(counter_analisys)] = key
			analisys_data[settings.data_fund_EUR_col+str(counter_analisys)] = fund_dict[key]["EUR"]
			analisys_data[settings.data_fund_ccy_col+str(counter_analisys)] = fund_dict[key]["ccy"]
			analisys_data[settings.data_fund_CCY_col+str(counter_analisys)] = fund_dict[key]["CCY"]
			counter_analisys = counter_analisys+1
		print("Data from book file loaded successfully!")

	def _get_system_names(self):
		bookname = self.bookname
		global set_systems
		systems = []
		book = load_workbook(self.directory+"/"+bookname+".xlsx")
		book_sheet = book.active
		curr = 3
		while curr<70000:
			if not book_sheet["A"+str(curr)].value == None:
				systems.append(str(book_sheet["A"+str(curr)].value))
			curr = curr + 1
		systems = list(set(systems))
		print(systems)
		return systems

	def _set_bookname(self, root):
		root.geometry("320x150")
		window = Frame(root, relief='sunken', bd=1, bg="#FFFFFF")
		window.pack(fill = "both", expand=True, padx=10, pady=10)

		def set_bookname():
			global set_bookname
			global default_settings
			bookname = field.get()
			self.bookname = field.get()
			default_settings = var1.get()

			try:
				bookname
				if str(bookname) == 'None' or int(bookname)<=0 or int(bookname)>999:
					messagebox.showwarning("Incorrect book name", "Please eneter valid bookname.")
					field.delete(0, END)
					return 0
			except ValueError:
				messagebox.showwarning("Incorrect book name", "Please eneter valid bookname.")
				field.delete(0, END)
				return 0

			files_ok = self.check_files()
			if files_ok:
				window.destroy()
				if default_settings == 0:
					self.get_system_names(root)
				else:
					global accr_sys_name, fund_sys_name
					self.settings.accr_sys_name = self.settings.d_accr_sys_name
					self.settings.fund_sys_name = self.settings.d_fund_sys_name
					self._create_analisys_file()
					quit()

		def quit():
			root.destroy()

		def toggle(var1):
			if var1.get() == 0:
				var1.set(1)
			else: var1.set(0)

		message1 = Label(window, text="Please enter a book name:", bg="#FFFFFF")
		message1.grid(row=1, column=0, padx=15, pady=15)

		field = ttk.Entry(window, width=10, justify=CENTER)
		field.grid(row=1, column=1, padx=5, pady=15)
		field.focus_set()

		var1 = IntVar()
		var1.set(1)
		t1 = Checkbutton(window, text="Use default settings", variable=var1, onvalue=1, offvalue=0, command=lambda:toggle(var1))
		t1.select()
		t1.grid(row=2, column=0)


		button1 = Button(window, text="Submit", command=set_bookname, width=10, bg="#DDDDDD", activebackground="#BBBBBB")
		button1.grid(row=3, column=0)

		button2 = Button(window, text="Quit", command=quit, width=10, bg="#DDDDDD", activebackground="#BBBBBB")
		button2.grid(row=3, column=1)

		root.mainloop()

	def _create_analisys_file(self):
		bookname = self.bookname
		analisys = load_workbook(self.directory+"/template.xlsx")
		analisys_data = analisys["data"]
		analisys_output = analisys["output"]

		self._load_data_from_external(analisys_data)
		self._load_data_from_internal(analisys_data, analisys_output)

		analisys.save(self.directory+"/"+bookname+"_internal_to_external.xlsx")
		print("Done!")

	def compare_book(self):
		root = Tk()
		root.title("Compare a book to external")
		self._set_bookname(root)
